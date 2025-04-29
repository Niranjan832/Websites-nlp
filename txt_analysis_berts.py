import os
import re
import pandas as pd
import requests
import torch
from bs4 import BeautifulSoup
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from openpyxl import load_workbook
from typing import List, Tuple

# Define paths using os.path.join for cross-platform compatibility
base_path = r"path"
input_path = os.path.join(base_path, "Output Data Structure.xlsx")
output_path = input_path

# Initialize BERT model and tokenizer
tokenizer = AutoTokenizer.from_pretrained('nlptown/bert-base-multilingual-uncased-sentiment')
model = AutoModelForSequenceClassification.from_pretrained('nlptown/bert-base-multilingual-uncased-sentiment')

def simple_word_tokenize(text: str) -> List[str]:
    """Simple word tokenizer that splits on whitespace and punctuation"""
    return re.findall(r'\b\w+\b', text.lower())

def simple_sentence_tokenize(text: str) -> List[str]:
    """Simple sentence tokenizer that splits on common sentence endings"""
    sentences = re.split(r'[.!?]+\s+(?=[A-Z])', text)
    return [s.strip() for s in sentences if s.strip()]

def extract_article_text(url: str) -> str:
    """Extract article text from URL using BeautifulSoup"""
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        title = soup.find('h1').get_text(strip=True) if soup.find('h1') else ''
        paragraphs = soup.find_all('p')
        content = ' '.join(p.get_text(strip=True) for p in paragraphs)
        return title + ' ' + content
    except Exception as e:
        print(f"Warning: Error extracting text from URL: {str(e)}")
        return ''

def get_bert_sentiment(text: str) -> Tuple[float, float, float]:
    """
    Calculate sentiment scores using BERT
    Returns: (positive_score, negative_score, compound_score)
    """
    try:
        # Truncate text to BERT's maximum length
        encoded = tokenizer(text, truncation=True, padding=True, return_tensors='pt')
        
        with torch.no_grad():
            output = model(**encoded)
            scores = torch.nn.functional.softmax(output.logits, dim=1)
            
        # BERT model returns scores from 1 to 5 stars
        # Convert to positive/negative/compound scores
        positive_score = float(scores[0][3:].sum())  # 4 and 5 stars
        negative_score = float(scores[0][:2].sum())  # 1 and 2 stars
        neutral_score = float(scores[0][2])          # 3 stars
        
        # Calculate compound score (-1 to 1 range)
        compound_score = (positive_score - negative_score) / (positive_score + negative_score + neutral_score)
        
        return positive_score, negative_score, compound_score
        
    except Exception as e:
        print(f"Warning: Error in BERT sentiment analysis: {str(e)}")
        return 0.0, 0.0, 0.0

def syllable_count(word: str) -> int:
    """Count syllables in a word"""
    word = word.lower()
    count = len(re.findall(r'[aeiouy]', word))
    if word.endswith(('es', 'ed')):
        count -= 1
    return max(1, count)

def calculate_scores(text: str) -> List[float]:
    """Calculate all text metrics including BERT sentiment"""
    try:
        words = simple_word_tokenize(text)
        cleaned_words = [word for word in words if word.isalnum()]
        sentences = simple_sentence_tokenize(text)

        if not cleaned_words or not sentences:
            return [0] * 13

        # Get BERT sentiment scores
        positive_score, negative_score, polarity_score = get_bert_sentiment(text)
        
        # Calculate subjectivity (ratio of opinionated content)
        subjectivity_score = abs(polarity_score)

        # Calculate readability metrics
        avg_sentence_length = len(cleaned_words) / len(sentences)
        complex_words = [word for word in cleaned_words if syllable_count(word) > 2]
        percentage_complex_words = len(complex_words) / len(cleaned_words)
        fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)

        # Calculate other text metrics
        word_count = len(cleaned_words)
        personal_pronouns = len(re.findall(r'\b(I|we|my|ours|us)\b', text, re.I))
        avg_word_length = sum(len(word) for word in cleaned_words) / len(cleaned_words)
        avg_syllables = sum(syllable_count(word) for word in cleaned_words) / len(cleaned_words)

        return [
            positive_score,
            negative_score,
            polarity_score,
            subjectivity_score,
            avg_sentence_length,
            percentage_complex_words,
            fog_index,
            avg_sentence_length,
            len(complex_words),
            word_count,
            avg_syllables,
            personal_pronouns,
            avg_word_length
        ]
    except Exception as e:
        print(f"Warning: Error calculating scores: {str(e)}")
        return [0] * 13

def main():
    try:
        # Check if input file exists
        if not os.path.exists(input_path):
            print(f"Error: Input Excel file not found at {input_path}")
            return
        
        # Read Excel file
        try:
            df = pd.read_excel(input_path)
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return
            
        results = []
        for index, row in df.iterrows():
            try:
                url = row['URL']
                article_text = extract_article_text(url)
                if not article_text:
                    results.append([row['URL_ID'], url] + [None] * 13)
                    continue
                    
                scores = calculate_scores(article_text)
                results.append([row['URL_ID'], url] + scores)
                print(f"Processed URL {index + 1}/{len(df)}: {url}")
            except Exception as e:
                print(f"Error processing URL {url}: {str(e)}")
                results.append([row['URL_ID'], url] + [None] * 13)
                
        # Write results back to Excel
        try:
            wb = load_workbook(output_path)
            sheet = wb.active
            
            for i, result in enumerate(results, start=2):
                for j, value in enumerate(result, start=1):
                    sheet.cell(row=i, column=j, value=value)
                    
            wb.save(output_path)
            print("Analysis completed successfully!")
        except Exception as e:
            print(f"Error writing results to Excel: {str(e)}")
            
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")

if __name__ == "__main__":
    main()
