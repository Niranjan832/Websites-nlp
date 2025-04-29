import os
import re
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Define paths using os.path.join for cross-platform compatibility
base_path = r"Path"
input_path = os.path.join(base_path, "Output Data Structure.xlsx")
output_path = input_path
stopwords_path = os.path.join(base_path, "StopWords")
master_dict_path = os.path.join(base_path, "MasterDictionary")

# Word lists content
POSITIVE_WORDS = """good
great
excellent
positive
outstanding
wonderful
amazing
awesome
superb
fantastic
brilliant
exceptional
marvelous
terrific
phenomenal
impressive
remarkable
splendid
fabulous
delightful
perfect
magnificent
extraordinary
superior
admirable
favorable
beneficial
effective
efficient
successful"""

NEGATIVE_WORDS = """bad
poor
negative
terrible
horrible
awful
unpleasant
disappointing
inferior
inadequate
unsatisfactory
unfavorable
deficient
defective
faulty
flawed
problematic
ineffective
inefficient
unsuccessful
fail
failure
worse
worst
mediocre
subpar
disappointing
frustrating
concerning
troublesome"""

STOPWORDS = """i
me
my
myself
we
our
ours
ourselves
you
your
yours
yourself
yourselves
he
him
his
himself
she
her
hers
herself
it
its
itself
they
them
their
theirs
themselves
what
which
who
whom
this
that
these
those
am
is
are
was
were
be
been
being
have
has
had
having
do
does
did
doing
would
should
could
might
must
shall
will
can
a
an
the
and
but
if
or
because
as
until
while
of
at
by
for
with
about
against
between
into
through
during
before
after
above
below
to
from
up
down
in
out
on
off
over
under
again
further
then
once
here
there
when
where
why
how"""

def create_word_lists():
    """Create necessary directories and word list files if they don't exist"""
    # Create directories
    os.makedirs(stopwords_path, exist_ok=True)
    os.makedirs(master_dict_path, exist_ok=True)
    
    # Create positive words file
    pos_file = os.path.join(master_dict_path, "positive-words.txt")
    if not os.path.exists(pos_file):
        with open(pos_file, 'w', encoding='utf-8') as f:
            f.write(POSITIVE_WORDS)
    
    # Create negative words file
    neg_file = os.path.join(master_dict_path, "negative-words.txt")
    if not os.path.exists(neg_file):
        with open(neg_file, 'w', encoding='utf-8') as f:
            f.write(NEGATIVE_WORDS)
    
    # Create stopwords file
    stop_file = os.path.join(stopwords_path, "stopwords.txt")
    if not os.path.exists(stop_file):
        with open(stop_file, 'w', encoding='utf-8') as f:
            f.write(STOPWORDS)

def simple_word_tokenize(text):
    """Simple word tokenizer that splits on whitespace and punctuation"""
    return re.findall(r'\b\w+\b', text.lower())

def simple_sentence_tokenize(text):
    """Simple sentence tokenizer that splits on common sentence endings"""
    sentences = re.split(r'[.!?]+\s+(?=[A-Z])', text)
    return [s.strip() for s in sentences if s.strip()]

def load_stopwords(path):
    try:
        stop_words = set()
        files = os.listdir(path)
        for file in files:
            file_path = os.path.join(path, file)
            with open(file_path, 'r', encoding='utf-8') as f:
                stop_words.update(f.read().splitlines())
        return stop_words
    except Exception as e:
        print(f"Warning: Error loading stopwords: {str(e)}. Using default stopwords.")
        return set(STOPWORDS.splitlines())

def load_master_dictionary(path):
    try:
        pos_path = os.path.join(path, 'positive-words.txt')
        neg_path = os.path.join(path, 'negative-words.txt')
        
        with open(pos_path, 'r', encoding='utf-8') as f:
            positive_words = set(f.read().splitlines())
        
        with open(neg_path, 'r', encoding='utf-8') as f:
            negative_words = set(f.read().splitlines())
            
        return positive_words, negative_words
    except Exception as e:
        print(f"Warning: Error loading master dictionary: {str(e)}. Using default word lists.")
        return set(POSITIVE_WORDS.splitlines()), set(NEGATIVE_WORDS.splitlines())

def extract_article_text(url):
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        title = soup.find('h1').get_text(strip=True) if soup.find('h1') else ''
        paragraphs = soup.find_all('p')
        content = ' '.join(p.get_text(strip=True) for p in paragraphs)
        return title + ' ' + content
    except Exception as e:
        print(f"Warning: Error extracting text from URL: {str(e)}")
        return ''

def syllable_count(word):
    word = word.lower()
    count = len(re.findall(r'[aeiouy]', word))
    if word.endswith(('es', 'ed')):
        count -= 1
    return max(1, count)

def avg_syllable_per_word(words):
    if not words:
        return 0
    return sum(syllable_count(word) for word in words) / len(words)

def calculate_scores(text):
    try:
        words = simple_word_tokenize(text)
        cleaned_words = [word for word in words if word.isalnum() and word not in stop_words]
        sentences = simple_sentence_tokenize(text)

        if not cleaned_words or not sentences:
            return [0] * 13

        positive_score = sum(1 for word in cleaned_words if word in positive_words)
        negative_score = sum(1 for word in cleaned_words if word in negative_words)
        polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
        subjectivity_score = (positive_score + negative_score) / (len(cleaned_words) + 0.000001)

        avg_sentence_length = len(cleaned_words) / len(sentences)
        complex_words = [word for word in cleaned_words if syllable_count(word) > 2]
        percentage_complex_words = len(complex_words) / len(cleaned_words)
        fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)

        word_count = len(cleaned_words)
        personal_pronouns = len(re.findall(r'\b(I|we|my|ours|us)\b', text, re.I))
        avg_word_length = sum(len(word) for word in cleaned_words) / len(cleaned_words)

        return [
            positive_score,
            negative_score,
            polarity_score,
            subjectivity_score,
            avg_sentence_length,
            percentage_complex_words,
            fog_index,
            avg_sentence_length,
            len(complex_words),
            word_count,
            avg_syllable_per_word(cleaned_words),
            personal_pronouns,
            avg_word_length
        ]
    except Exception as e:
        print(f"Warning: Error calculating scores: {str(e)}")
        return [0] * 13

def main():
    try:
        # Create word lists if they don't exist
        create_word_lists()
        
        # Load required resources
        global stop_words, positive_words, negative_words
        stop_words = load_stopwords(stopwords_path)
        positive_words, negative_words = load_master_dictionary(master_dict_path)
        
        # Check if input file exists
        if not os.path.exists(input_path):
            print(f"Error: Input Excel file not found at {input_path}")
            return
        
        # Read Excel file
        try:
            df = pd.read_excel(input_path)
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return
            
        results = []
        for index, row in df.iterrows():
            try:
                url = row['URL']
                article_text = extract_article_text(url)
                if not article_text:
                    results.append([row['URL_ID'], url] + [None] * 13)
                    continue
                    
                scores = calculate_scores(article_text)
                results.append([row['URL_ID'], url] + scores)
                print(f"Processed URL {index + 1}/{len(df)}: {url}")
            except Exception as e:
                print(f"Error processing URL {url}: {str(e)}")
                results.append([row['URL_ID'], url] + [None] * 13)
                
        # Write results back to Excel
        try:
            wb = load_workbook(output_path)
            sheet = wb.active
            
            for i, result in enumerate(results, start=2):
                for j, value in enumerate(result, start=1):
                    sheet.cell(row=i, column=j, value=value)
                    
            wb.save(output_path)
            print("Analysis completed successfully!")
        except Exception as e:
            print(f"Error writing results to Excel: {str(e)}")
            
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")

if __name__ == "__main__":
    main()
